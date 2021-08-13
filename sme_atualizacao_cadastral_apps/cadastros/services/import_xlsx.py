from django.db import transaction
from openpyxl import load_workbook

from sme_atualizacao_cadastral_apps.cadastros.helpers import salvar_log
from sme_atualizacao_cadastral_apps.cadastros.models import BaseCadastro


def import_xlsx(planilha):
    nome_arquivo = str(planilha.arquivo.name.split("/")[-1])
    if planilha.extraido:
        salvar_log(
            arquivo=nome_arquivo,
            status=False,
            msg_retorno='Esse registro já foi processdo anteriormente.'
        )
    filepath = planilha.arquivo.path
    wb = load_workbook(filepath)
    ws = wb.worksheets[0]

    nova_base = []
    for i in range(2, ws.max_row + 1):
        cpf = ws.cell(row=i, column=1)
        situacao = ws.cell(row=i, column=2)

        cadastro = dict(
            cpf=str(cpf.value),
            situacao=situacao.value
        )

        if cpf and situacao:
            obj = BaseCadastro(**cadastro)
            nova_base.append(obj)
    try:
        with transaction.atomic():
            BaseCadastro.objects.all().delete()
            BaseCadastro.objects.bulk_create(nova_base)
            planilha.extraido = True
            planilha.save()
            salvar_log(
                arquivo=nome_arquivo,
                status=True,
                msg_retorno='Planilha processada com sucesso.'
            )
    except ValueError as e:
        salvar_log(
            arquivo=nome_arquivo,
            status=False,
            msg_retorno=str(e)
        )
